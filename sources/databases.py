import sources.logins as lg 
from sqlite3 import connect
from IPython.display import display
import pandas as pd
import requests
from io import BytesIO
import openpyxl

class databases():

	def __init__(self,uname):
		self.update_type = [
							'ExFM',
							'Master List',
							'Both',
							'Price',
							'Installation',
							'Customers',
							'All',
			]
		if uname not in ('Le Quang Thong','Admin'):

			self.update_type = ['ExFM',]
		
			
		self.uname = uname	
		# self.update_type.sort()

	def select_db_name(self):
		print('\n')
		files = lg.file_filter(end_with='db',printer = True)
		print ('   0  |  CREATE NEW DATABASE')
		while True:
			try:
				ind = int(input('Select Database: '))
			except:
				print('Only accept by number')
				ind =''
			if ind == 0:
				while True:
					db_name = str(input('Database name: '))
					
					#check extension
					if not db_name.endswith('.db'):
						db_name = db_name + '.db'
					
					# check exists	
					if db_name in files.values():
						print('Data base already exists. Choose another name.')

					
					else:
						break
						
			if ind == 0:
				print(f'Conncect Success to {db_name}')
				conn = connect(db_name)
				break
			if ind <= len(files):
				conn = connect(files[ind])
				print(f'Connect Success to {files[ind]}')
				db_name = files[ind]
				break
			else:
				print(f'Index must be less than {len(files)}')
		
		try:
			conn.close()
			return db_name
		except Exception as e:
			print(e)


	def update_db(self,db_name):
		ans = str(input('Update Database Y/[N]:') or 'n')
		
		if ans.upper()=='N':
			conn = connect(db_name)
			print(f'\nUsing {db_name}')
		else:
			update_type = self.update_type
			conn = connect(db_name)

			#Table Header
			print(f'\n{"_"*50}')
			print(f'{"|  No.|  Update Type": <49}|')
			print(f'|{"_"*48}|')

			#Contents
			for i in range(1,1+len(update_type)):
				print(f'|{i: >3}  |  {update_type[i-1]: <40}|')
			print(f'|{"_"*48}|') #bottom border
			
			#select type of update
			while True:
				ind = str(input('Select Function by Index: '))
				try:
					ind = int(ind)
				except:
					print('Only accept number')
					continue
				if 0<ind<=len(update_type):
					
					u_type = update_type[ind-1]
					break
				else:
					print(f'Input number must be less than {len(update_type)}\n')

			#start update by type
			if u_type =='All':u_type = update_type # convert all u_type
			
			if 'ExFM' in u_type:
				print('\nSelect file ExFM:\n')
				con = lg.file_select(folder_name='files',start_with = 'SearchResult')
				try:
					df = pd.read_excel(con,sheet_name=None)
					c = df['Consolidated'].drop(['OWNERSHIP'],axis=1)
					c.to_sql('consolidated',conn,index=False,if_exists='replace')
					(df['PF-Code']).to_sql('pf_code',conn,index=False,if_exists='replace')
					(df['CD-Code']).to_sql('cd_code',conn,index=False,if_exists='replace')
					(df['R-Code']).to_sql('repair_code',conn,index=False,if_exists='replace')
					(df['Parts']).to_sql('parts',conn,index=False,if_exists='replace')
					print(f'\nConsolidated file stored in {db_name}')
				except Exception as e:
					print(f'can not find file {con} and import as consolidated')
					print(e)

				try:
					# google sheet
					spreadsheetId = '1bT4W0CiLVD_B_ddRcVkS3MEXbSjvmGb4'
					url = "https://docs.google.com/spreadsheets/export?exportFormat=xlsx&id=" + spreadsheetId
					res = requests.get(url)
					data = BytesIO(res.content)
					xlsx = openpyxl.load_workbook(filename=data)
					for name in xlsx.sheetnames:
						values = pd.read_excel(data, sheet_name=name)
						values.to_sql(name,conn,index=False,if_exists='replace')
					print(f'\nDatabase file stored in {db_name}')
				except Exception as e:
					print(f'can not connect with Google Sheet')
					print(e)

			if 'Master List' in u_type:
				print('\nSelect file Master List:\n')
				filename = lg.file_select(folder_name='files',end_with='.xlsm')
				try:
					# m_list
					m_list = pd.read_excel(filename,sheet_name='Master List',skiprows=range(1,2))
					new_header = m_list.iloc[0] #grab the first row for the header
					m_list = m_list[1:] #take the data less the header row
					m_list.columns = new_header

					
					m_list.to_sql('new_ml',conn,index=False,if_exists='replace')
				except Exception as e:
					print(e)

			if 'both' in u_type.lower(): # ExFM & Master List
				print('\nSelect file ExFM:\n')
				con = lg.file_select(folder_name='files',start_with = 'SearchResult')
				try:
					df = pd.read_excel(con,sheet_name=None)
					c = df['Consolidated'].drop(['OWNERSHIP'],axis=1)
					c.to_sql('consolidated',conn,index=False,if_exists='replace')
					(df['PF-Code']).to_sql('pf_code',conn,index=False,if_exists='replace')
					(df['CD-Code']).to_sql('cd_code',conn,index=False,if_exists='replace')
					(df['R-Code']).to_sql('repair_code',conn,index=False,if_exists='replace')
					(df['Parts']).to_sql('parts',conn,index=False,if_exists='replace')
					print(f'\nConsolidated file stored in {db_name}')

					# google sheet
					spreadsheetId = '1bT4W0CiLVD_B_ddRcVkS3MEXbSjvmGb4'
					url = "https://docs.google.com/spreadsheets/export?exportFormat=xlsx&id=" + spreadsheetId
					res = requests.get(url)
					data = BytesIO(res.content)
					xlsx = openpyxl.load_workbook(filename=data)
					for name in xlsx.sheetnames:
						values = pd.read_excel(data, sheet_name=name)
						values.to_sql(name,conn,index=False,if_exists='replace')
					print(f'\nDatabase file stored in {db_name}')
				except Exception as e:
					print(f'can not find file {con} and import as consolidated')
					print(e)

				print('\nSelect file Master List:\n')
				filename = lg.file_select(folder_name='files',end_with='.xlsm')
				try:
					# m_list
					m_list = pd.read_excel(filename,sheet_name='Master List',skiprows=range(1,2))
					new_header = m_list.iloc[0] #grab the first row for the header
					m_list = m_list[1:] #take the data less the header row
					m_list.columns = new_header

					
					m_list.to_sql('new_ml',conn,index=False,if_exists='replace')
				except Exception as e:
					print(e)
					
			if 'Google Sheet Data' in u_type:
				# try:
				spreadsheetId = '1bT4W0CiLVD_B_ddRcVkS3MEXbSjvmGb4'
				url = "https://docs.google.com/spreadsheets/export?exportFormat=xlsx&id=" + spreadsheetId
				res = requests.get(url)
				data = BytesIO(res.content)
				xlsx = openpyxl.load_workbook(filename=data)
				for name in xlsx.sheetnames:
					values = pd.read_excel(data, sheet_name=name)
					values.to_sql(name,conn,index=False,if_exists='replace')
				print(f'\nDatabase file stored in {db_name}')
				# except:
				# 	print(f'can not find file Google Sheet Data')

			if 'Price' in u_type:
				print('\nSelect File Price:\n')
				try:
					price_file = lg.file_select(contains='rice',folder_name = 'files')
					xl = pd.ExcelFile(price_file)
					i=1
					for sh in xl.sheet_names:  # see all sheet names
						print(i,'  |  ',sh)
						i+=1
					sh_index = int(input('Select file index: '))-1
					pr = pd.read_excel(price_file,sheet_name = xl.sheet_names[sh_index])
					pr.to_sql('prices',conn,index = False, if_exists = 'replace')

				except Exception as e:
					print(e)
			if 'Installation' in u_type:
				print('\nSelect file Installation:\n')
				try:
					ins = lg.file_select(folder_name = 'files',start_with = 'EQP')
					df1 = pd.read_excel(ins,sheet_name=None)
					(df1['csvdata']).to_sql('install',conn,index=False,if_exists='replace')
					print(f'\nInstallation file stored in {db_name}')
				except:
					print(f'can not find file {ins} and import as installation data')
			if 'Customers' in u_type:
				print('\nSelect file ACC_TBL_EXP:\n')
				try:
					ins = lg.file_select(folder_name = 'files',start_with='ACC')
					df1 = pd.read_excel(ins,sheet_name=None)
					(df1['csvdata']).to_sql('acc_tbl_exp',conn,index=False,if_exists='replace')
					print(f'\nCustomer file stored in {db_name}')
				except:
					print(f'can not find file {ins} and import as installation data')
			# else:
			# 	print(f'Select wrong {u_type}')

			# display(pd.read_sql('SELECT name from sqlite_master where type= "table";',conn))

		return conn


